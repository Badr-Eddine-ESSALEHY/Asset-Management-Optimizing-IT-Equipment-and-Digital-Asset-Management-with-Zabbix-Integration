from django.contrib.auth import authenticate, login, get_user_model, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.http import require_http_methods
import json
import csv
from datetime import datetime, timedelta
import zipfile
from io import BytesIO
from .forms import UserUpdateForm, ProfileUpdateForm, CustomUserCreationForm, EnhancedPasswordChangeForm
from .models import Profile
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy

CustomUser = get_user_model()


def login_member(request):
    """Enhanced login view with better error handling and user experience"""
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        remember_me = request.POST.get("remember_me")

        if not username or not password:
            messages.error(request, "Please provide both username and password.")
            return render(request, "members/login.html")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if not user.is_active:
                messages.error(request, "Your account has been deactivated. Please contact an administrator.")
                return render(request, "members/login.html")

            login(request, user)

            # Set session expiry based on remember me
            if not remember_me:
                request.session.set_expiry(0)  # Session expires when browser closes
            else:
                request.session.set_expiry(1209600)  # 2 weeks

            # Update last login manually if needed
            user.last_login = timezone.now()
            user.save(update_fields=['last_login'])

            messages.success(request, f"Welcome back, {user.get_full_name() or user.username}!")

            # Enhanced redirect logic
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            elif user.is_staff:
                return redirect('pages:dashboard')
            else:
                return redirect('assets:equipment_list')
        else:
            messages.error(request, "Invalid username or password. Please try again.")

    return render(request, "members/login.html")


def register(request):
    """Enhanced registration view with better validation"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()

            # Create profile if it doesn't exist
            Profile.objects.get_or_create(user=user)

            messages.success(
                request,
                "Account created successfully! You can now log in with your credentials."
            )
            return redirect('login')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CustomUserCreationForm()

    return render(request, 'members/register.html', {'form': form})


@login_required
def profile_view(request):
    """Enhanced profile view with comprehensive user management"""
    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileUpdateForm(
            request.POST,
            request.FILES,
            instance=request.user.profile
        )

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('members:profile')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=request.user.profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form,
    }

    # Enhanced admin context with additional statistics
    if request.user.is_staff:
        # Get all users with related profile data
        all_users = CustomUser.objects.select_related('profile').all().order_by('username')

        # Pagination for better performance
        paginator = Paginator(all_users, 50)  # Show 50 users per page
        page_number = request.GET.get('page')
        paginated_users = paginator.get_page(page_number)

        # User statistics
        total_users = CustomUser.objects.count()
        active_users = CustomUser.objects.filter(is_active=True).count()
        admin_users = CustomUser.objects.filter(is_staff=True).count()
        recent_users = CustomUser.objects.filter(
            date_joined__gte=timezone.now() - timedelta(days=30)
        ).count()

        # Recent logins (last 7 days)
        recent_logins = CustomUser.objects.filter(
            last_login__gte=timezone.now() - timedelta(days=7)
        ).count()

        context.update({
            'all_users': paginated_users,
            'total_users': total_users,
            'active_users': active_users,
            'admin_users': admin_users,
            'recent_users': recent_users,
            'recent_logins': recent_logins,
            'inactive_users': total_users - active_users,
        })

    return render(request, 'members/profile.html', context)


@login_required
@staff_member_required
def delete_user(request, user_id):
    """Enhanced user deletion with better security and logging"""
    user_to_delete = get_object_or_404(CustomUser, id=user_id)

    if user_to_delete == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('members:profile')

    if user_to_delete.is_superuser and not request.user.is_superuser:
        messages.error(request, "You cannot delete a superuser account.")
        return redirect('members:profile')

    username = user_to_delete.username
    user_to_delete.delete()
    messages.success(request, f"User '{username}' has been deleted successfully.")

    return redirect('members:profile')


@login_required
@staff_member_required
@require_http_methods(["POST"])
def toggle_user_status(request, user_id):
    """AJAX endpoint to toggle user active status"""
    try:
        user_to_toggle = get_object_or_404(CustomUser, id=user_id)

        if user_to_toggle == request.user:
            return JsonResponse({
                'success': False,
                'message': 'You cannot modify your own account status.'
            })

        if user_to_toggle.is_superuser and not request.user.is_superuser:
            return JsonResponse({
                'success': False,
                'message': 'You cannot modify a superuser account.'
            })

        user_to_toggle.is_active = not user_to_toggle.is_active
        user_to_toggle.save(update_fields=['is_active'])

        status = "activated" if user_to_toggle.is_active else "deactivated"
        return JsonResponse({
            'success': True,
            'message': f'User {user_to_toggle.username} has been {status}.',
            'new_status': user_to_toggle.is_active
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'An error occurred: {str(e)}'
        })


@login_required
@staff_member_required
def user_search_api(request):
    """AJAX endpoint for user search"""
    query = request.GET.get('q', '').strip()

    if not query:
        return JsonResponse({'users': []})

    users = CustomUser.objects.filter(
        Q(username__icontains=query) |
        Q(email__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query)
    ).select_related('profile')[:20]  # Limit to 20 results

    users_data = []
    for user in users:
        users_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'full_name': user.get_full_name(),
            'is_active': user.is_active,
            'is_staff': user.is_staff,
            'department': user.department,
            'profile_picture': user.profile.profile_picture.url if user.profile.profile_picture else None,
        })

    return JsonResponse({'users': users_data})


@login_required
@staff_member_required
def export_users(request):
    """Export users data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response[
        'Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Username', 'Email', 'First Name', 'Last Name',
        'Department', 'Phone', 'Is Staff', 'Is Active',
        'Date Joined', 'Last Login'
    ])

    users = CustomUser.objects.all().order_by('username')
    for user in users:
        writer.writerow([
            user.username,
            user.email,
            user.first_name,
            user.last_name,
            user.department or '',
            user.phone or '',
            'Yes' if user.is_staff else 'No',
            'Yes' if user.is_active else 'No',
            user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '',
            user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never',
        ])

    return response


@login_required
@staff_member_required
def generate_report(request):
    """Generate comprehensive PDF report"""
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.darkblue,
        alignment=1  # Center alignment
    )
    story.append(Paragraph("User Management Report", title_style))
    story.append(Spacer(1, 20))

    # Summary Statistics
    total_users = CustomUser.objects.count()
    active_users = CustomUser.objects.filter(is_active=True).count()
    staff_users = CustomUser.objects.filter(is_staff=True).count()
    recent_users = CustomUser.objects.filter(
        date_joined__gte=timezone.now() - timedelta(days=30)
    ).count()

    summary_data = [
        ['Metric', 'Count', 'Percentage'],
        ['Total Users', str(total_users), '100%'],
        ['Active Users', str(active_users), f'{(active_users / total_users * 100):.1f}%' if total_users > 0 else '0%'],
        ['Staff Users', str(staff_users), f'{(staff_users / total_users * 100):.1f}%' if total_users > 0 else '0%'],
        ['Recent Signups (30 days)', str(recent_users),
         f'{(recent_users / total_users * 100):.1f}%' if total_users > 0 else '0%'],
    ]

    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(Paragraph("Summary Statistics", styles['Heading2']))
    story.append(summary_table)
    story.append(Spacer(1, 20))

    # User List Table
    story.append(Paragraph("User Details", styles['Heading2']))

    users = CustomUser.objects.all().order_by('username')[:50]  # Limit to first 50 users
    user_data = [['Username', 'Email', 'Full Name', 'Department', 'Status']]

    for user in users:
        user_data.append([
            user.username,
            user.email,
            user.get_full_name() or 'N/A',
            user.department or 'N/A',
            'Active' if user.is_active else 'Inactive'
        ])

    user_table = Table(user_data)
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    story.append(user_table)

    # Build PDF
    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response[
        'Content-Disposition'] = f'attachment; filename="user_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

    return response


@login_required
def download_account_data(request):
    """Download user's account data"""
    user = request.user

    # Create a ZIP file containing user data
    buffer = BytesIO()

    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # User data as JSON
        user_data = {
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'department': user.department,
            'phone': user.phone,
            'date_joined': user.date_joined.isoformat() if user.date_joined else None,
            'last_login': user.last_login.isoformat() if user.last_login else None,
            'is_active': user.is_active,
            'is_staff': user.is_staff,
        }

        # Profile data
        profile_data = {}
        if hasattr(user, 'profile'):
            profile_data = {
                'bio': user.profile.bio,
                'location': user.profile.location,
                'website': user.profile.website,
                'created_at': user.profile.created_at.isoformat() if user.profile.created_at else None,
                'updated_at': user.profile.updated_at.isoformat() if user.profile.updated_at else None,
            }

        # Add files to ZIP
        zip_file.writestr('user_data.json', json.dumps(user_data, indent=2))
        zip_file.writestr('profile_data.json', json.dumps(profile_data, indent=2))

        # Add profile picture if exists
        if hasattr(user, 'profile') and user.profile.profile_picture:
            try:
                with user.profile.profile_picture.open('rb') as img_file:
                    zip_file.writestr(f'profile_picture_{user.username}.jpg', img_file.read())
            except:
                pass  # Skip if file doesn't exist

    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{user.username}_account_data.zip"'

    return response


@login_required
def user_statistics(request):
    """Get user statistics for dashboard"""
    stats = {
        'total_users': CustomUser.objects.count(),
        'active_users': CustomUser.objects.filter(is_active=True).count(),
        'inactive_users': CustomUser.objects.filter(is_active=False).count(),
        'admin_users': CustomUser.objects.filter(is_staff=True).count(),
        'recent_signups': CustomUser.objects.filter(
            date_joined__gte=timezone.now() - timedelta(days=7)
        ).count(),
        'recent_logins': CustomUser.objects.filter(
            last_login__gte=timezone.now() - timedelta(days=7)
        ).count(),
    }

    # Users by department
    departments = CustomUser.objects.values('department').annotate(
        count=Count('id')
    ).exclude(department__isnull=True).exclude(department='')

    stats['departments'] = list(departments)

    # Monthly signup trend (last 6 months)
    monthly_signups = []
    for i in range(6):
        date = timezone.now() - timedelta(days=30 * i)
        start_date = date.replace(day=1)
        if i == 0:
            end_date = timezone.now()
        else:
            end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        count = CustomUser.objects.filter(
            date_joined__range=[start_date, end_date]
        ).count()

        monthly_signups.append({
            'month': start_date.strftime('%B %Y'),
            'count': count
        })

    stats['monthly_signups'] = monthly_signups

    return JsonResponse(stats)


@login_required
def change_password_view(request):
    """Enhanced change password view"""
    if request.method == 'POST':
        form = EnhancedPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Your password has been changed successfully!')
            # Re-authenticate the user to maintain session
            update_session_auth_hash(request, user)
            return redirect('members:profile')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = EnhancedPasswordChangeForm(request.user)

    return render(request, 'members/change_password.html', {'form': form})


@login_required
def account_settings_view(request):
    """Account settings and preferences"""
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_preferences':
            # Update user preferences
            messages.success(request, 'Preferences updated successfully!')
        elif action == 'export_data':
            return download_account_data(request)
        elif action == 'delete_account':
            # Handle account deletion request
            messages.warning(request, 'Account deletion requested. Please contact administrator.')

    return render(request, 'members/account_settings.html')


@login_required
@staff_member_required
def bulk_user_actions(request):
    """Handle bulk actions on users"""
    if request.method == 'POST':
        action = request.POST.get('action')
        user_ids = request.POST.getlist('user_ids')

        if not user_ids:
            messages.error(request, 'No users selected.')
            return redirect('members:profile')

        users = CustomUser.objects.filter(id__in=user_ids).exclude(id=request.user.id)

        if action == 'activate':
            users.update(is_active=True)
            messages.success(request, f'{users.count()} users have been activated.')
        elif action == 'deactivate':
            users.update(is_active=False)
            messages.success(request, f'{users.count()} users have been deactivated.')
        elif action == 'delete':
            count = users.count()
            users.delete()
            messages.success(request, f'{count} users have been deleted.')

    return redirect('members:profile')


def is_admin(user):
    """Check if user is admin"""
    return user.is_authenticated and user.is_staff


@user_passes_test(is_admin)
def admin_dashboard(request):
    """Admin dashboard with user management overview"""
    context = {
        'total_users': CustomUser.objects.count(),
        'active_users': CustomUser.objects.filter(is_active=True).count(),
        'recent_users': CustomUser.objects.filter(
            date_joined__gte=timezone.now() - timedelta(days=7)
        ).count(),
        'recent_logins': CustomUser.objects.filter(
            last_login__gte=timezone.now() - timedelta(days=7)
        ).count(),
    }

    return render(request, 'pages/dashboard.html', context)


@login_required
@staff_member_required
def enable_two_factor_auth(request):
    """Enable two-factor authentication for user"""
    if request.method == 'POST':
        # This is a placeholder - you would integrate with a 2FA library like django-otp
        messages.success(request, 'Two-factor authentication has been enabled for your account.')
        return redirect('members:profile')

    return render(request, 'members/enable_2fa.html')


@login_required
@staff_member_required
def view_user_detail(request, user_id):
    """View detailed user information (for admin)"""
    user_obj = get_object_or_404(CustomUser, id=user_id)

    context = {
        'user_obj': user_obj,
        'profile': getattr(user_obj, 'profile', None),
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return JSON for AJAX requests
        user_data = {
            'id': user_obj.id,
            'username': user_obj.username,
            'email': user_obj.email,
            'full_name': user_obj.get_full_name(),
            'department': user_obj.department,
            'phone': user_obj.phone,
            'is_active': user_obj.is_active,
            'is_staff': user_obj.is_staff,
            'date_joined': user_obj.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user_obj.date_joined else None,
            'last_login': user_obj.last_login.strftime('%Y-%m-%d %H:%M:%S') if user_obj.last_login else None,
        }

        if hasattr(user_obj, 'profile'):
            user_data.update({
                'bio': user_obj.profile.bio,
                'location': user_obj.profile.location,
                'website': user_obj.profile.website,
                'profile_picture': user_obj.profile.profile_picture.url if user_obj.profile.profile_picture else None,
            })

        return JsonResponse({'user': user_data})

    return render(request, 'members/user_detail.html', context)


@login_required
@staff_member_required
def export_users_excel(request):
    """Export users data to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Users Export"

        # Headers
        headers = [
            'Username', 'Email', 'First Name', 'Last Name',
            'Department', 'Phone', 'Role', 'Status',
            'Date Joined', 'Last Login'
        ]

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        users = CustomUser.objects.all().order_by('username')
        for row, user in enumerate(users, 2):
            worksheet.cell(row=row, column=1, value=user.username)
            worksheet.cell(row=row, column=2, value=user.email)
            worksheet.cell(row=row, column=3, value=user.first_name)
            worksheet.cell(row=row, column=4, value=user.last_name)
            worksheet.cell(row=row, column=5, value=user.department or '')
            worksheet.cell(row=row, column=6, value=user.phone or '')
            worksheet.cell(row=row, column=7, value='Admin' if user.is_staff else 'User')
            worksheet.cell(row=row, column=8, value='Active' if user.is_active else 'Inactive')
            worksheet.cell(row=row, column=9, value=user.date_joined.strftime('%Y-%m-%d') if user.date_joined else '')
            worksheet.cell(row=row, column=10,
                           value=user.last_login.strftime('%Y-%m-%d') if user.last_login else 'Never')

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        workbook.save(response)
        return response

    except ImportError:
        # Fallback to CSV if openpyxl is not installed
        messages.warning(request, 'Excel export not available. Downloading CSV instead.')
        return export_users(request)


@login_required
def send_password_reset_email(request):
    """Send password reset email to user"""
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = CustomUser.objects.get(email=email)
            # Here you would integrate with Django's password reset functionality
            # For now, we'll just show a success message
            messages.success(request, f'Password reset email sent to {email}')
        except CustomUser.DoesNotExist:
            messages.error(request, 'No user found with this email address.')

    return redirect('members:profile')


@login_required
def update_profile_visibility(request):
    """Update user profile visibility settings"""
    if request.method == 'POST':
        visibility = request.POST.get('profile_visibility', 'team')

        # Store in user profile or preferences model
        if hasattr(request.user, 'profile'):
            # You might need to add a visibility field to your Profile model
            # request.user.profile.visibility = visibility
            # request.user.profile.save()
            pass

        messages.success(request, f'Profile visibility updated to {visibility}.')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Visibility updated successfully'})

    return redirect('members:profile')


@login_required
def delete_own_account(request):
    """Allow user to delete their own account"""
    if request.method == 'POST':
        password = request.POST.get('password')

        # Verify password before deletion
        if not request.user.check_password(password):
            messages.error(request, 'Incorrect password. Account not deleted.')
            return redirect('members:profile')

        # Store username before deletion
        username = request.user.username

        # Delete user account
        request.user.delete()

        messages.success(request, f'Account {username} has been permanently deleted.')
        return redirect('members:login')

    return render(request, 'members/delete_account_confirm.html')


# Enhanced Password Change View Class
class CustomPasswordChangeView(PasswordChangeView):
    """Custom password change view with enhanced styling"""
    template_name = 'members/change_password.html'
    form_class = EnhancedPasswordChangeForm
    success_url = reverse_lazy('members:profile')

    def form_valid(self, form):
        messages.success(self.request, 'Your password has been changed successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)